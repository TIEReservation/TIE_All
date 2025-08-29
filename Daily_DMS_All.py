import time
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
import openpyxl
from datetime import datetime
import gspread
from oauth2client.service_account import ServiceAccountCredentials

# Dictionary of properties and their hotel IDs
PROPERTIES = {
        "EdenBeachResort": "30357",
        "Villa Shakti": "27724",
        "Le Pondy Beachside": "27723",
        "Le Royce Villa": "27722",
        "Le Poshe Suite": "27721",
        "Le Poshe Luxury": "27720",
        "Le Poshe Beach View": "27719",
        "La Villa Heritage": "27711",
        "La Tamara suite": "27710",
        "La Tamara Luxury": "27709",
        "La Paradise Residency": "27707",
        "La Paradise Luxury": "27706",
        "La Antilia Luxury": "27704",
        "La Millionaire Resort": "31550",
        "Le Park Resort": "32470"
}

def setup_driver(chrome_profile_path):
    """Set up Chrome WebDriver with the specified user profile."""
    chrome_options = Options()
    chrome_options.add_argument(f"user-data-dir={chrome_profile_path}")
    chrome_options.add_argument("profile-directory=Profile 20")
    chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
    chrome_options.add_argument("--log-level=3")
    driver = webdriver.Chrome(options=chrome_options)
    return driver

def login_to_stayflexi(chrome_profile_path, property_name, hotel_id):
    """Login to StayFlexi, navigate to the property dashboard, and access reservations."""
    driver = setup_driver(chrome_profile_path)
    wait = WebDriverWait(driver, 20)
    bookings = []
    
    try:
        print(f"🔹 Opening StayFlexi login page for {property_name} (Hotel ID: {hotel_id})...")
        driver.get("https://app.stayflexi.com/auth/login")
        
        try:
            print("🔹 Attempting to login...")
            email_field = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@type='text']")))
            email_field.clear()
            email_field.send_keys("gayathri.tie@gmail.com")
            
            login_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(text(),'Sign In')]")))
            login_button.click()
            
            password_field = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@type='password']")))
            password_field.send_keys("Alliswell@2025")
            
            login_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(text(),'Sign In')]")))
            login_button.click()
            print("✅ Logged in successfully")
            time.sleep(5)
            
        except Exception:
            print("🔹 Already logged in, skipping login step.")

        print(f"🔹 Navigating to {property_name} Dashboard...")
        dashboard_button = wait.until(EC.element_to_be_clickable(
            (By.XPATH, f"//a[@href='/dashboard?hotelId={hotel_id}']")))
        dashboard_button.click()
        print(f"✅ Opened {property_name} Dashboard")
        
        time.sleep(3)
        driver.switch_to.window(driver.window_handles[-1])
        print("🔄 Switched to new tab")
        
        print("🔹 Clicking on Reservations button...")
        reservations_button = wait.until(EC.element_to_be_clickable(
            (By.XPATH, "//button[contains(text(), 'Reservations')]")))
        reservations_button.click()
        print("✅ Clicked on Reservations button")

        """print("🔹 Clicking on New Bookings...")
        new_bookings_button = wait.until(EC.element_to_be_clickable(
            (By.XPATH, "//*[@id=\"kt_content\"]/div/div/div[1]/div/div[1]/div[3]/div[1]/div")))
        new_bookings_button.click()
        print("✅ Clicked on New Bookings")"""
        
        print(f"✅ Logged in and ready to fetch bookings for {property_name}")
        bookings = fetch_and_display_bookings(driver, wait, hotel_id)
        
        for booking in bookings:
            fetch_folio_details(driver, wait, booking, hotel_id)
            update_sheets(booking, driver, property_name)
        
    except Exception as e:
        print(f"❌ Error for {property_name}: {str(e)}")
    finally:
        return driver, bookings

def fetch_and_display_bookings(driver, wait, hotel_id):
    """Fetch and display all booking information entries."""
    print("🔹 Fetching all booking information entries...")
    bookings = []

    time.sleep(8)

    try:
        booking_cards = wait.until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div.MuiCollapse-root.MuiCollapse-vertical.MuiCollapse-hidden")))
        print(f"📋 Found {len(booking_cards)} booking entries using MuiCollapse-hidden")
    except Exception as e:
        print(f"⚠️ Could not find booking entries with MuiCollapse-hidden: {str(e)}")
        try:
            booking_cards = wait.until(
                EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div.MuiAccordionSummary-content.Mui-expanded.MuiAccordionSummary-contentGutters")))
            print(f"📋 Found {len(booking_cards)} booking entries using MuiAccordionSummary-expanded")
        except Exception as e:
            print(f"⚠️ Could not find booking entries with MuiAccordionSummary-expanded: {str(e)}")
            booking_cards = []

    for i, card in enumerate(booking_cards):
        print(f"\n🔖 Booking #{i+1}:")
        try:
            if "MuiCollapse-hidden" in card.get_attribute("class"):
                print("  ℹ️ Element is collapsed, attempting to expand...")
                accordion_button = card.find_element(By.XPATH, "./preceding-sibling::div[contains(@class, 'MuiAccordionSummary-root')]")
                driver.execute_script("arguments[0].scrollIntoView(); arguments[0].click();", accordion_button)
                time.sleep(2)

            accordion = card.find_element(By.XPATH, "./ancestor::div[contains(@class, 'MuiAccordion-root')]")
            summary_content = accordion.find_element(By.CSS_SELECTOR, "div.MuiAccordionSummary-content")
            raw_text = summary_content.text.strip()

            print(f"  📄 Raw text: {raw_text[:200]}{'...' if len(raw_text) > 200 else ''}")

            booking_data = extract_booking_data_from_text(raw_text, hotel_id)
            display_booking_data(booking_data)
            bookings.append(booking_data)

        except Exception as e:
            print(f"  ❌ Error processing booking: {str(e)}")

    if not booking_cards:
        match_patterns_on_page(driver, hotel_id)

    return bookings

def extract_booking_data_from_text(text, hotel_id):
    """Extract booking information including room number and type from text."""
    booking_data = {
        'name': None,
        'booking_id': None,
        'phone': None,
        'booking_period': None,
        'booking_source': None,
        'total_without_taxes': None,
        'total_tax_amount': None,
        'total_with_taxes': None,
        'payment_made': None,
        'balance_due': None,
        'room_number': 'N/A',
        'room_type': 'N/A',
        'rate_plan': 'N/A',
        'adults_children_infant': 'N/A'
    }

    lines = text.split('\n')

    if lines and not re.search(r'SFBOOKING|Rs\.|CONFIRMED|ON_HOLD|Mar| - |[0-9]', lines[0]):
        booking_data['name'] = lines[0].strip()

    booking_id_match = re.search(rf'SFBOOKING_{hotel_id}_\d+', text)
    if booking_id_match:
        booking_data['booking_id'] = booking_id_match.group(0)

    for line in lines:
        line = line.strip()
        if re.match(r'NA|(\+\d{1,3}\s*)?[\d\s()-]{8,}', line):
            booking_data['phone'] = line
            break

    date_pattern = r'(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{1,2},\s+\d{4}\s+\d{1,2}:\d{2}\s+(?:AM|PM)\s+-\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{1,2},\s+\d{4}\s+\d{1,2}:\d{2}\s+(?:AM|PM)'
    date_match = re.search(date_pattern, text)
    if date_match:
        booking_data['booking_period'] = date_match.group(0)
    else:
        for i in range(len(lines) - 1):
            if " - " in lines[i] and re.search(r'(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)', lines[i]):
                booking_data['booking_period'] = f"{lines[i].strip()} - {lines[i+1].strip()}"
                break

    room_pattern = r'(\d+)\s*\(\s*([^)]+)\s*\)'
    room_match = re.search(room_pattern, text)
    if room_match:
        booking_data['room_number'] = room_match.group(1).strip()
        booking_data['room_type'] = room_match.group(2).strip()

    return booking_data

def fetch_folio_details(driver, wait, booking, hotel_id):
    """Navigate to the folio page and fetch financial details, Rate Plan, and Adults/Children/Infant."""
    try:
        if booking['booking_id']:
            folio_url = f"https://app.stayflexi.com/folio/{booking['booking_id']}?hotelId={hotel_id}"
            print(f"🔹 Navigating to folio page for {booking['booking_id']}...")
            driver.get(folio_url)
            time.sleep(5)

            try:
                expand_button = wait.until(EC.element_to_be_clickable(
                    (By.CSS_SELECTOR, ".MuiAccordionSummary-expandIconWrapper.css-1fx8m19")))
                driver.execute_script("arguments[0].scrollIntoView(); arguments[0].click();", expand_button)
                print("  ✅ Clicked down arrow button using CSS selector on View Folio page")
            except Exception as e:
                print(f"  ⚠️ Could not click down arrow using CSS selector: {str(e)}")

            time.sleep(2)

            try:
                booking_source_elem = wait.until(EC.presence_of_element_located(
                    (By.XPATH, "//div[@class='sourceName' and contains(text(), 'BOOKING.COM')]")))
                booking['booking_source'] = booking_source_elem.text.strip()
                print(f"📋 Booking Source: {booking['booking_source']}")
            except Exception as e:
                print(f"⚠️ Could not fetch booking source: {str(e)}")

            try:
                rate_plan_elem = wait.until(EC.presence_of_element_located(
                    (By.XPATH, "//*[@id='panel1a-content']/div/div/div[1]/div/div[8]/div/div[2]")))
                booking['rate_plan'] = rate_plan_elem.text.strip()
                print(f"📋 Rate Plan: {booking['rate_plan']}")
            except Exception as e:
                booking['rate_plan'] = 'N/A'
                print(f"⚠️ Could not fetch Rate Plan: {str(e)}")

            try:
                adults_children_infant_elem = wait.until(EC.presence_of_element_located(
                    (By.XPATH, "//*[@id='panel1a-content']/div/div/div[2]/div/div[8]/div/div[2]")))
                booking['adults_children_infant'] = adults_children_infant_elem.text.strip()
                print(f"📋 Adults/Children/Infant: {booking['adults_children_infant']}")
            except Exception as e:
                booking['adults_children_infant'] = 'N/A'
                print(f"⚠️ Could not fetch Adults/Children/Infant: {str(e)}")

            try:
                financial_section = wait.until(EC.presence_of_element_located(
                    (By.XPATH, "//*[@id='kt_content']/div/div/div[1]/div/div[2]/div/div[2]/div")))
                financial_text = financial_section.text.strip().split('\n')

                for i, line in enumerate(financial_text):
                    if "Total without taxes" in line:
                        booking['total_without_taxes'] = re.sub(r'(INR|Rs\.)\s*', '', financial_text[i + 1].strip())
                    elif "Total tax amount" in line:
                        booking['total_tax_amount'] = re.sub(r'(INR|Rs\.)\s*', '', financial_text[i + 1].strip())
                    elif "Total with taxes and fees" in line:
                        booking['total_with_taxes'] = re.sub(r'(INR|Rs\.)\s*', '', financial_text[i + 1].strip())
                    elif "Payment made" in line:
                        booking['payment_made'] = re.sub(r'(INR|Rs\.)\s*', '', financial_text[i + 1].strip())
                    elif "Balance due" in line:
                        booking['balance_due'] = re.sub(r'(INR|Rs\.)\s*', '', financial_text[i + 1].strip())

                print(f"💰 Total without taxes: {booking['total_without_taxes']}")
                print(f"💰 Total tax amount: {booking['total_tax_amount']}")
                print(f"💰 Total with taxes: {booking['total_with_taxes']}")
                print(f"💰 Payment made: {booking['payment_made']}")
                print(f"💰 Balance due: {booking['balance_due']}")
            except Exception as e:
                print(f"⚠️ Could not fetch financial details: {str(e)}")
                
            print(f"✅ Successfully fetched folio details for {booking['booking_id']}")
        else:
            print("⚠️ No booking ID found, skipping folio fetch")
            
    except Exception as e:
        print(f"❌ Error fetching folio details: {str(e)}")

def match_patterns_on_page(driver, hotel_id):
    """Look for booking patterns directly on page using JavaScript."""
    print("🔍 Executing JavaScript to find booking patterns...")
    
    js_find_bookings = f"""
    const elements = [];
    const nodeIterator = document.createNodeIterator(
        document.body,
        NodeFilter.SHOW_TEXT,
        {{ acceptNode: function(node) {{ 
            return node.textContent.includes('SFBOOKING_{hotel_id}') ? NodeFilter.FILTER_ACCEPT : NodeFilter.FILTER_REJECT;
        }} }}
    );
    
    let node;
    while(node = nodeIterator.nextNode()) {{
        let parent = node.parentElement;
        let depth = 0;
        while(parent && depth < 5) {{
            if (parent.offsetWidth > 100 && parent.offsetHeight > 50) {{
                elements.push({{
                    text: parent.innerText,
                    html: parent.outerHTML
                }});
                break;
            }}
            parent = parent.parentElement;
            depth++;
        }}
    }}
    return elements;
    """
    
    booking_elements = driver.execute_script(js_find_bookings)
    
    if booking_elements:
        print(f"🎯 Found {len(booking_elements)} booking elements using JavaScript")
        for i, elem in enumerate(booking_elements[:3]):
            print(f"\n🔖 Booking Element #{i+1}:")
            text = elem.get('text', '')
            print(f"📄 Raw text: {text[:200]}...")
            soup = BeautifulSoup(elem.get('html', ''), 'html.parser')
            booking_data = extract_booking_data_from_text(text, hotel_id)
            display_booking_data(booking_data)
    else:
        print("⚠️ No booking elements found using JavaScript")
        print("📸 Taking screenshots for debugging...")
        driver.execute_script("window.scrollTo(0, 0)")
        driver.save_screenshot(f"stayflexi_top_{hotel_id}.png")
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
        driver.save_screenshot(f"stayflexi_bottom_{hotel_id}.png")
        print("✅ Saved screenshots")

def display_booking_data(booking_data):
    """Display the extracted booking data in a formatted way."""
    has_data = False
    if booking_data['name']:
        print(f"👤 Name: {booking_data['name']}")
        has_data = True
    if booking_data['booking_id']:
        print(f"🔑 Booking ID: {booking_data['booking_id']}")
        has_data = True
    if booking_data['phone']:
        print(f"📞 Phone: {booking_data['phone']}")
        has_data = True
    if booking_data['booking_period']:
        print(f"🗓️ Booking Period: {booking_data['booking_period']}")
        has_data = True
    if booking_data['room_number']:
        print(f"🏠 Room Number: {booking_data['room_number']}")
        has_data = True
    if booking_data['room_type']:
        print(f"🏠 Room Type: {booking_data['room_type']}")
        has_data = True
    if booking_data['rate_plan']:
        print(f"📋 Rate Plan: {booking_data['rate_plan']}")
        has_data = True
    if booking_data['adults_children_infant']:
        print(f"👨‍👩‍👧 Adults/Children/Infant: {booking_data['adults_children_infant']}")
        has_data = True
    
    if not has_data:
        print("ℹ️ No booking data found")

def update_sheets(booking, driver, property_name):
    """Update both local Excel and Google Sheets with booking details if not already present."""
    property_columns_excel = [
        "Report Date", "Booking Date", "Booking Id", "Booking Source", 
        "Guest Name", "Guest Phone", "Check In", "Check Out", "Total with taxes", 
        "Payment Made", "Adults/Children/Infant", "Room Number", "Total without taxes", 
        "Total tax amount", "Room Type", "Rate Plan"
    ]
    all_properties_columns_excel = [
        "Property", "Report Date", "Booking Date", "Booking Id", "Booking Source", 
        "Guest Name", "Guest Phone", "Check In", "Check Out", "Total with taxes", 
        "Payment Made", "Adults/Children/Infant", "Room Number", "Total without taxes", 
        "Total tax amount", "Room Type", "Rate Plan"
    ]

    property_columns_gsheet = [
        "Report Date", "Booking Date", "Booking Id", "Booking Source", 
        "Guest Name", "Guest Phone", "Check In", "Check Out", "Total with taxes", 
        "Payment Made", "Adults/Children/Infant", "Room Number", "Total without taxes", 
        "Total tax amount", "Room Type", "Rate Plan"
    ]
    all_properties_columns_gsheet = [
        "Property", "Report Date", "Booking Date", "Booking Id", "Booking Source", 
        "Guest Name", "Guest Phone", "Check In", "Check Out", "Total with taxes", 
        "Payment Made", "Adults/Children/Infant", "Room Number", "Total without taxes", 
        "Total tax amount", "Room Type", "Rate Plan"
    ]

    report_date = datetime.now().strftime("%Y-%m-%d")
    booking_date = report_date
    check_in = booking['booking_period'].split(' - ')[0] if booking['booking_period'] else ""
    check_out = booking['booking_period'].split(' - ')[1] if booking['booking_period'] else ""

    if check_in:
        try:
            check_in_date = datetime.strptime(check_in, "%b %d, %Y %I:%M %p")
            check_in = check_in_date.strftime("%Y-%m-%d")
        except ValueError as e:
            print(f"⚠️ Could not parse Check In date '{check_in}': {str(e)}")
            check_in = ""

    if check_out:
        try:
            check_out_date = datetime.strptime(check_out, "%b %d, %Y %I:%M %p")
            check_out = check_out_date.strftime("%Y-%m-%d")
        except ValueError as e:
            print(f"⚠️ Could not parse Check Out date '{check_out}': {str(e)}")
            check_out = ""

    property_row_excel = [
        report_date, booking_date, booking.get('booking_id', ''), booking.get('booking_source', ''),
        booking.get('name', ''), booking.get('phone', ''), check_in, check_out,
        booking.get('total_with_taxes', ''), booking.get('payment_made', ''), 
        booking.get('adults_children_infant', ''), booking.get('room_number', ''), 
        booking.get('total_without_taxes', ''), booking.get('total_tax_amount', ''),
        booking.get('room_type', ''), booking.get('rate_plan', '')
    ]

    all_properties_row_excel = [
        property_name, report_date, booking_date, booking.get('booking_id', ''),
        booking.get('booking_source', ''), booking.get('name', ''), booking.get('phone', ''),
        check_in, check_out, booking.get('total_with_taxes', ''), booking.get('payment_made', ''),
        booking.get('adults_children_infant', ''), booking.get('room_number', ''), 
        booking.get('total_without_taxes', ''), booking.get('total_tax_amount', ''),
        booking.get('room_type', ''), booking.get('rate_plan', '')
    ]

    property_row_gsheet = [
        report_date, booking_date, booking.get('booking_id', ''), booking.get('booking_source', ''),
        booking.get('name', ''), booking.get('phone', ''), check_in, check_out,
        booking.get('total_with_taxes', ''), booking.get('payment_made', ''),
        booking.get('adults_children_infant', ''), booking.get('room_number', ''), 
        booking.get('total_without_taxes', ''), booking.get('total_tax_amount', ''),
        booking.get('room_type', ''), booking.get('rate_plan', '')
    ]

    all_properties_row_gsheet = [
        property_name, report_date, booking_date, booking.get('booking_id', ''),
        booking.get('booking_source', ''), booking.get('name', ''), booking.get('phone', ''),
        check_in, check_out, booking.get('total_with_taxes', ''), booking.get('payment_made', ''),
        booking.get('adults_children_infant', ''), booking.get('room_number', ''), 
        booking.get('total_without_taxes', ''), booking.get('total_tax_amount', ''),
        booking.get('room_type', ''), booking.get('rate_plan', '')
    ]

    # Load or create Excel workbook
    try:
        workbook = openpyxl.load_workbook("DMS_DetailedSheet.xlsx")
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])

    # Property-specific sheet
    if property_name not in workbook.sheetnames:
        property_sheet = workbook.create_sheet(property_name)
        property_sheet.append(property_columns_excel)
    else:
        property_sheet = workbook[property_name]

    property_exists = False
    for row in range(2, property_sheet.max_row + 1):
        existing_id = property_sheet.cell(row=row, column=3).value
        if existing_id == booking.get('booking_id'):
            property_exists = True
            print(f"ℹ️ Booking ID {booking.get('booking_id')} already exists in {property_name} sheet (local)")
            break

    if not property_exists:
        property_sheet.append(property_row_excel)
        print(f"✅ Updated {property_name} sheet (local) for {booking['name']} with Booking ID {booking.get('booking_id')}")

    # All Properties sheet
    if "All Properties" not in workbook.sheetnames:
        all_properties_sheet = workbook.create_sheet("All Properties")
        all_properties_sheet.append(all_properties_columns_excel)
    else:
        all_properties_sheet = workbook["All Properties"]

    all_properties_exists = False
    for row in range(2, all_properties_sheet.max_row + 1):
        existing_id = all_properties_sheet.cell(row=row, column=4).value
        if existing_id == booking.get('booking_id'):
            all_properties_exists = True
            print(f"ℹ️ Booking ID {booking.get('booking_id')} already exists in All Properties sheet (local)")
            break

    if not all_properties_exists:
        all_properties_sheet.append(all_properties_row_excel)
        print(f"✅ Updated All Properties sheet (local) for {booking['name']} with Booking ID {booking.get('booking_id')}")

    workbook.save("DMS_DetailedSheet.xlsx")
    print("✅ Local workbook saved successfully")

    # Update Google Sheets
    spreadsheet_id = '1L8THT5kqoa0-pgo55J9R6f351dnz_9kCTBs9-oyFT9s'
    try:
        scope = [
            'https://spreadsheets.google.com/feeds',
            'https://www.googleapis.com/auth/drive'
        ]
        credentials = ServiceAccountCredentials.from_json_keyfile_name('credentials.json', scope)
        client = gspread.authorize(credentials)
        sheet = client.open_by_key(spreadsheet_id)

        # Property-specific worksheet
        try:
            property_worksheet = sheet.worksheet(property_name)
            existing_data = property_worksheet.get_all_values()
            existing_ids = [row[2] for row in existing_data[1:] if len(row) > 2]
        except gspread.exceptions.WorksheetNotFound:
            property_worksheet = sheet.add_worksheet(title=property_name, rows=1000, cols=20)
            property_worksheet.append_row(property_columns_gsheet)
            existing_ids = []
            print(f"Created new worksheet: {property_name} (Google Sheet)")

        if booking.get('booking_id') not in existing_ids:
            property_worksheet.append_row(property_row_gsheet)
            print(f"✅ Updated {property_name} sheet (Google Sheet) for {booking['name']} with Booking ID {booking.get('booking_id')}")
        else:
            print(f"ℹ️ Booking ID {booking.get('booking_id')} already exists in {property_name} sheet (Google Sheet)")

        # All Properties worksheet
        try:
            all_properties_worksheet = sheet.worksheet("All Properties")
            existing_data = all_properties_worksheet.get_all_values()
            existing_composite_keys = [f"{row[0]}_{row[3]}" for row in existing_data[1:] if len(row) > 3]
        except gspread.exceptions.WorksheetNotFound:
            all_properties_worksheet = sheet.add_worksheet(title="All Properties", rows=1000, cols=20)
            all_properties_worksheet.append_row(all_properties_columns_gsheet)
            existing_composite_keys = []
            print("Created new worksheet: All Properties (Google Sheet)")

        composite_key = f"{property_name}_{booking.get('booking_id')}"
        if composite_key not in existing_composite_keys:
            all_properties_worksheet.append_row(all_properties_row_gsheet)
            print(f"✅ Updated All Properties sheet (Google Sheet) for {booking['name']} with Booking ID {booking.get('booking_id')}")
        else:
            print(f"ℹ️ Booking ID {booking.get('booking_id')} already exists in All Properties sheet (Google Sheet)")

        print(f"✅ Data successfully saved to Google Sheet: https://docs.google.com/spreadsheets/d/{spreadsheet_id}")

    except Exception as e:
        print(f"⚠️ Error connecting to Google Sheets API: {str(e)}")
        print("Data was saved locally but not to Google Sheets.")

def process_all_properties(chrome_profile_path):
    """Process all properties sequentially without waiting for user input."""
    for property_name, hotel_id in PROPERTIES.items():
        print(f"\n{'='*50}\nProcessing {property_name} (Hotel ID: {hotel_id})\n{'='*50}")
        driver, _ = login_to_stayflexi(chrome_profile_path, property_name, hotel_id)
        driver.quit()
        print(f"✅ Finished processing {property_name}, moving to the next property...")

if __name__ == "__main__":
    chrome_profile_path = r"C:\Users\somas\AppData\Local\Google\Chrome\User Data\Default"
    process_all_properties(chrome_profile_path)